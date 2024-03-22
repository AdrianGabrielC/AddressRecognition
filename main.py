import csv
import json
import re

import openpyxl
import requests
import spacy
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
from spacy.tokens import DocBin
from openpyxl.utils import get_column_letter

### PHASE 1. Identify addresses ###
def scrape_page(url):
    try:
        response = requests.get(url)
        html_content = response.content
        soup = BeautifulSoup(html_content, "html.parser")
        text_content = soup.get_text(" ")
        text_content = ' '.join(text_content.split())
        return text_content
    except:
        return None

def predict_addresses():
    wb = openpyxl.Workbook()
    ws = wb.active

    model = spacy.load('PhaseOne/model/model-best')
    with open('websites_list.csv', mode='r', newline='', encoding='utf-8') as file:
        csv_reader = csv.reader(file)
        #row_count = 0
        excel_row = 1
        flag = False
        skip_flag = False
        for index, row in enumerate(reversed(list(csv_reader)[:10])):
            website_url = f"https://{row[0]}"
            txt = scrape_page(website_url)
            if txt is not None:
                doc = model(txt)
                for ent in doc.ents:
                    ws.cell(row=excel_row, column=1).value = website_url
                    ws.cell(row=excel_row, column=2).value = sanitize(ent.text)
                    excel_row = excel_row + 1
                    print("Excel row:", excel_row, "Website:", website_url, "Address:", ent.text, ent.label_)

        wb.save('prediction_result_test.xlsx')

def sanitize(text):
    allowed_characters = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789 ,.?'!#"
    pattern = f"[^{re.escape(allowed_characters)}]"
    return re.sub(pattern, "", text)

def clean_full_address():
    def truncate_string(input_string, max_length):
        if len(input_string) > max_length:
            return input_string[:max_length]
        return input_string

    # Load the Excel workbook
    wb = openpyxl.load_workbook('prediction_result_test.xlsx')
    sheet = wb.active

    # Iterate over rows in column B and truncate strings
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
        cell_value = row[0].value
        if cell_value:
            row[0].value = truncate_string(cell_value, 93)

    # Iterate through rows and remove those where the last word in column B doesn't contain a digit
    rows_to_delete = []
    for row in sheet.iter_rows(min_row=2, max_col=2, max_row=sheet.max_row):
        cell_value = row[1].value
        if cell_value:
            last_word = cell_value.split()[-1]
            if not any(char.isdigit() for char in last_word):
                rows_to_delete.append(row)

    # Delete rows from the worksheet
    for row in rows_to_delete:
        sheet.delete_rows(row[0].row)

    # Save the modified workbook
    wb.save('prediction_result_test_clean.xlsx')



### PHASE TWO. Identify street number, street name, city, region and ZIP ###

# Create spacy dataset
def create_train_data_spacy_address():
    nlp = spacy.blank('en')
    db = DocBin()
    for index in range(26):
        data = json.load(open(f'./address_dataset/annotations ({index}).json'))
        text = data['annotations'][0][0]
        entities = data['annotations'][0][1]['entities']
        doc = nlp.make_doc(text)
        ents = []
        for start, end, label in entities:
            print(start, end, label)
            try:
                span = doc.char_span(start, end, label=label, alignment_mode='strict')
            except:
                continue
            if span is not None:
                ents.append(span)
        try:
            doc.ents = ents
            db.add(doc)
        except:
            pass

    data = json.load(open(f'./address_dataset/usa_annotations.json'))
    for annotation in data['annotations']:
        if annotation is not None:
            text = annotation[0]
            entities = annotation[1]['entities']
            doc = nlp.make_doc(text)
            ents = []
            for start, end, label in entities:
                print(start, end, label)
                try:
                    span = doc.char_span(start, end, label=label, alignment_mode='strict')
                except:
                    continue
                if span is not None:
                    ents.append(span)
            try:
                doc.ents = ents
                db.add(doc)
            except:
                pass
    db.to_disk('./address_dataset/train_address_data.spacy')

# Predict address components
def predict_address_components():
    data = []
    wb = openpyxl.load_workbook('./address_dataset/prediction_result_combined_v4.xlsx')
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        website, text = row[:2]
        data.append([website, text])

    model = spacy.load('./address_dataset/model-best')
    result = []
    for row in data:
        doc = model(row[1])
        ents = []
        for ent in doc.ents:
            ents.append([ent.text, ent.label_])
        result.append([row[0], doc.text, ents])
    return result

def write_to_excel(results):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws['A1'] = 'Website'
    ws['B1'] = 'Full Address'
    ws['C1'] = 'Street Number'
    ws['D1'] = 'Street Name'
    ws['E1'] = 'City'
    ws['F1'] = 'Region'
    ws['G1'] = 'ZIP'

    for row_idx, result in enumerate(results, start=2):
        website = result[0]
        full_address = result[1]
        ents = result[2]

        # Write website and full address to columns A and B
        ws.cell(row=row_idx, column=1, value=website)
        ws.cell(row=row_idx, column=2, value=full_address)

        # Extract information from ents
        street_number = ''
        street_name = ''
        city = ''
        region = ''
        zip_code = ''
        for item in ents:
            if item[1] == 'STREET_NUMBER':
                street_number = item[0]
            elif item[1] == 'STREET_NAME':
                street_name = item[0]
            elif item[1] == 'CITY':
                city = item[0]
            elif item[1] == 'REGION':
                region = item[0]
            elif item[1] == 'ZIP':
                zip_code = item[0]

        # Write extracted information to columns C-G
        ws.cell(row=row_idx, column=3, value=street_number)
        ws.cell(row=row_idx, column=4, value=street_name)
        ws.cell(row=row_idx, column=5, value=city)
        ws.cell(row=row_idx, column=6, value=region)
        ws.cell(row=row_idx, column=7, value=zip_code)

        # Save the Excel file
    wb.save('./address_dataset/final_addresses.xlsx')


